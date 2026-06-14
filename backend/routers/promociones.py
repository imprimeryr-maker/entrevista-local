import base64
import io
import openpyxl

from fastapi import APIRouter, HTTPException

from backend.database import (
    get_promociones as db_get_promociones,
    get_proyectos as db_get_proyectos,
    create_promocion as db_create_promocion,
    get_promociones_by_mes as db_get_promociones_by_mes,
    get_meses_promociones as db_get_meses_promociones,
    update_promocion as db_update_promocion,
    delete_promocion as db_delete_promocion,
)
from backend.models import PromocionCreate, PromocionUpdate, AnalizarExcelRequest
from backend.services.ai_service import analizar_excel_promociones

router = APIRouter(prefix="/api/promociones", tags=["promociones"])


@router.get("")
def listar_promociones():
    return db_get_promociones()


@router.get("/meses")
def listar_meses():
    return db_get_meses_promociones()


@router.get("/por-mes/{mes}")
def listar_por_mes(mes: str):
    return db_get_promociones_by_mes(mes)


@router.post("")
def crear_promocion(body: PromocionCreate):
    return db_create_promocion(
        proyecto_id=body.proyecto_id,
        nombre_proyecto=body.nombre_proyecto,
        mes=body.mes,
        descripcion_promocion=body.descripcion_promocion,
        archivo_original=body.archivo_original,
    )


@router.put("/{promocion_id}")
def actualizar_promocion(promocion_id: str, body: PromocionUpdate):
    kwargs = {k: v for k, v in body.model_dump().items() if v is not None}
    if kwargs:
        db_update_promocion(promocion_id, **kwargs)
    return {"ok": True}


@router.delete("/{promocion_id}")
def eliminar_promocion(promocion_id: str):
    db_delete_promocion(promocion_id)
    return {"ok": True}


@router.post("/analizar-excel")
def analizar_excel(body: AnalizarExcelRequest):
    try:
        file_bytes = base64.b64decode(body.contenido_base64)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        hojas_texto = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            filas = []
            if headers:
                header_str = " | ".join(str(h) if h is not None else "" for h in headers)
                filas.append(f"ENCABEZADOS: {header_str}")
            for row in rows_iter:
                if any(cell is not None for cell in row):
                    fila_detalles = []
                    for i, cell in enumerate(row):
                        header_name = str(headers[i]) if headers and i < len(headers) else f"Col{i+1}"
                        val = str(cell) if cell is not None else ""
                        fila_detalles.append(f"{header_name}: {val}")
                    filas.append(" - ".join(fila_detalles))
            if filas:
                hojas_texto.append(f"=== Hoja: {sheet_name} ===\n" + "\n".join(filas))

        texto_completo = "\n\n".join(hojas_texto)
        if len(texto_completo) < 20:
            raise HTTPException(status_code=400, detail="El archivo Excel está vacío o no tiene datos legibles")

        resultados = analizar_excel_promociones(texto_completo, body.nombre_mes)
        if resultados is None:
            raise HTTPException(status_code=500, detail="Error de conexión con el proveedor AI")

        proyectos = db_get_proyectos()
        proyectos_dict = {p["nombre"].lower().strip(): p for p in proyectos}

        items_con_match = []
        for item in resultados:
            nombre_proyecto = item.get("nombre_proyecto", "").strip()
            if not nombre_proyecto:
                continue
            nombre_lower = nombre_proyecto.lower().strip()
            match = proyectos_dict.get(nombre_lower)
            if not match:
                for p_nombre, p in proyectos_dict.items():
                    if nombre_lower in p_nombre or p_nombre in nombre_lower:
                        match = p
                        break
            items_con_match.append({
                **item,
                "match": match["nombre"] if match else None,
                "proyecto_id": match["id"] if match else None,
            })

        return {"items": items_con_match}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al procesar Excel: {str(e)}")
